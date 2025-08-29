"""Takes excel, csv, tsv or txt files retrieved from pubmed, and saves each sheet as csv file
    Doesn't do evaluation (where is the data?) at this stage, just provides the columns in the tracking file so that other 
    programs can do this, and then data_convert can act on them if populated.
    Based on the original data_convert, which combined these functions.
"""

from fileinput import filename
import logging
import pandas as pd
import os
from openpyxl import load_workbook
import xlrd
import csv
import re
import argparse
from akg import AKGException, akg_logging_config
from tracking import create_tracking, load_tracking, save_tracking, create_empty_tracking_store, add_to_tracking, tracking_entry
import sys

def process_excel_file(file_path)->pd.DataFrame:
    """loads excel files into dataframes
    """
    # tracking DataFrame
    tdf = create_empty_tracking_store()

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        output_dir = os.path.dirname(file_path)
        
        for sheet_name in wb.sheetnames:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            new_file = process_dataframe(df, sheet_name, output_dir, file_path)
            tdf = add_to_tracking(tdf, new_file)

    except Exception as e:
        logging.error(f"Failed to process excel file: {file_path}: {str(e)}")

    return tdf

def process_old_file(file_path)->pd.DataFrame:
    """loads older-style excel files (.xls) into dataframes
    """
    # tracking DataFrame
    tdf = create_empty_tracking_store()

    try:
        wb = xlrd.open_workbook(file_path)
        output_dir = os.path.dirname(file_path)
        for sheet in wb.sheets():
            logging.info(f"Processing sheet: {sheet.name}")
            headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
            data = [
                [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                for row in range(1, sheet.nrows)
            ]
            df = pd.DataFrame(data, columns=headers)
            
            new_file = process_dataframe(df, sheet.name, output_dir, file_path)
            tdf = add_to_tracking(tdf, new_file)

    except Exception as e:
        logging.error(f"Failed to process 'old' file: {file_path}: {str(e)}")

    return tdf

def process_csv_file(file_path:str)->pd.DataFrame:
    """loads csv, tsc or txt files and prepares them to be inputs to the AKG

        Parameters:
            file_path:str   The file to process
        Returns:
            Information about the added output files (if any) in a form suitable for adding to the tracking data (a DataFrame)
    """
    # The processing creates files that need to be added to the tracking, which can't be done inside the loop.
# this is the pattern for adding some:
#    added_files = pd.DataFrame({col: pd.Series(dtype=dt) for col, dt in tracking_col_names.items()})
#    df = pd.concat([df,added_files],ignore_index=True)

    # tracking DataFrame
    tdf = create_empty_tracking_store()

    output_dir = os.path.dirname(file_path)
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    # Try reading with different settings
    for encoding in ['utf-8', 'iso-8859-1', 'latin1']:
        for delim in ['\t', ',', ';']:  # prioritize tab
            try:
                df = pd.read_csv(file_path, delimiter=delim, encoding=encoding, on_bad_lines='warn')
                if not df.empty:
                    new_file = process_dataframe(df, file_name, output_dir, file_path, input_delimiter=delim)
                    return add_to_tracking(tdf, new_file)
            except Exception as e:
                logging.error(f"Failed to read {file_path} with delimiter '{delim}' and encoding '{encoding}': {str(e)}")

    # If all attempts fail, try reading as plain text
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        logging.info(f"File {file_path} couldn't be processed as CSV. Content preview:")
        logging.info(content[:100])  # Print first 100 characters
    except Exception as e:
        logging.error(f"Failed to read {file_path} as text: {str(e)}")

    return tdf


def process_dataframe(df, sheet_name, output_dir, file_path, input_delimiter='\t')->pd.DataFrame:
    """processes dataframes to assess if the data relates to gene expression - looks for "log fold change" or similar
    in column titles
    """
    # tracking DataFrame. There should be a maximum of one entry in the returned value because this function works on a single dataframe
    tdf = create_empty_tracking_store()

    df.columns = df.columns.astype(str)
    
        # try to fix some odd characters
#        sheet_name = sheet_name.encode(encoding="ascii",errors="backslashreplace")
        # remove characters not appropriate for filenames
    replacement_chars  = {" " : "",
                            "<" : "lessthan",
                            ">" : "morethan",
                            "." : "",
                            ":" : "",
                            "/" : "",
                            "?" : "",
                            "*" : "",
                            "&" : "" }
    for old, new in replacement_chars.items():
        sheet_name = sheet_name.replace(old, new)
    new_filestub = f'split_{sheet_name}'
    new_filename = new_filestub+'.csv'
    output_file = os.path.join(output_dir, new_filename)

    if os.path.exists(output_file):
        original_filename = os.path.splitext(os.path.basename(file_path))[0]
        new_filename = f"{new_filestub}_{original_filename}.csv"
        output_file = os.path.join(output_dir, new_filename)

    if input_delimiter == '\t':
        df.to_csv(output_file, index=False, sep=',')
    else:
        df.to_csv(output_file, index=False)
    logging.info(f"Saved {sheet_name} as CSV: {output_file}")

    # assume the pmid is the last component of the output dir
    pmid = os.path.basename(output_dir)
    # create a new tracking entry
    new_entry = tracking_entry(1,output_dir,pmid,new_filename, False, True, file_path, False, False, '', 0, '', '', '','', 0, 0,False,'')

    tdf = add_to_tracking(tdf, new_entry)

    return tdf

def process_supp_data_folder(data_folder:str, tracking_file_path:str ):
    """ 
    function process_supp_data_folder

    Walks through all immediate PMID-named subdirectories of data_folder, processes files found according to their extension
    Excludes subdirectories which are marked as 'exclude' in the CSV file defined in tracking_file_path
    *May* change contents of article_file_path at this level in future, doesn't do this at present

    Parameters:
        data_folder:str
        tracking_file_path:str # must be a full path

    Returns:
        None

    Raises:
        No direct exception handling/raising in this code    

    """
    df = load_tracking(tracking_file_path)

# The processing creates files that need to be added to the tracking, which can't be done inside the loop.
    tdf  = create_empty_tracking_store()
    local_tdf = create_empty_tracking_store()
# TODO: remove loop iteration, do sthg more pythonic
    for index, row in df.iterrows():
        # data_split only works on step 0 files, the raw data that was downloaded
        if row['step'] == 0 and not row['excl']:
            root = row['path']
            file = row['file']
            file_path = os.path.join(root, file)
            # never process files that we wrote out on a previous iteration
            if file.lower().startswith('expdata_') or file.lower().startswith('split_'):
                logging.info(f"Skipping file: {file_path}")
                continue
            logging.info(f"Processing file: {file_path}")
            if file.lower().endswith('.xlsx'):
                local_tdf = process_excel_file(file_path)
            elif file.lower().endswith('.xls'):
                local_tdf = process_old_file(file_path)
            elif file.lower().endswith(('.csv', '.tsv', '.txt')):
                local_tdf = process_csv_file(file_path)
            tdf = add_to_tracking(tdf,local_tdf)
            # flag the source data as excluded, just for completeness.
            # see the check above. This means that if you rerun data_split, the same file will not be processed twice unless you
            # change the 'excl' flag back to False.
            df.loc[int(index),'excl'] = True
    logging.info(f'Finished processing files in {data_folder}, found {len(tdf)} new files to add to tracking')

    # now the loop has finished add the accumulated tracking info into the main one
    df = add_to_tracking(df, tdf)

    # write out the updated information (should have the new files we just wrote out)
    save_tracking(df, tracking_file_path)


if __name__ == '__main__':

    command_line_str = ' '.join(sys.argv)

    # manage the command line options
    parser = argparse.ArgumentParser(description='Splits downloaded supplementary data in .xlsx files (and others) into .csv files, one per table')
    parser.add_argument('-i','--input_dir', default='data', help='Destination top-level directory for input data files (output files also written here)')
    parser.add_argument('-t','--tracking_file', default='akg_tracking.xlsx', help='Tracking file name. This file is created in the top-level directory.')
    parser.add_argument('-l', '--log', default='data_split.log', help='Log file name. This file is created in the top-level directory.')

    # argparse populates an object using parse_args
    # extract its members into a dict and from there into variables if used in more than one place
    config = vars(parser.parse_args())

    main_dir = config['input_dir']

    if not os.path.isdir(main_dir):
        raise AKGException(f"data_convert: data directory {main_dir} must exist") 

    akg_logging_config(os.path.join(main_dir, config['log']))
    logging.info(f'Starting data_split in {main_dir}')
    logging.info(f"Program executed with command: {command_line_str}")

    # create the tracking file    
    tracking_file = config['tracking_file']

    logging.info(f'Processing directory {os.path.realpath(main_dir)}: creating tracking file {tracking_file} here if necessary')
    tracking_file = os.path.join(main_dir, tracking_file)

    log_file = config['log']
    log_file = os.path.join(main_dir, log_file)

    if not os.path.exists(tracking_file):
        create_tracking(main_dir, tracking_file)

    supp_data_folder = os.path.join(main_dir,"supp_data")
    process_supp_data_folder(supp_data_folder, tracking_file)

